from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
import csv
import re
from typing import Any
import unicodedata

from openpyxl import load_workbook

from app.models.integration import IntegrationMatchStatus


SOURCE_ROW_KEY = "__source_row__"


@dataclass
class ProcessResult:
    """Resultado do processamento de arquivo: estatísticas + lista de matches."""
    total_registros: int
    total_matches: int
    total_divergencias: int
    matches: list[dict[str, Any]]


# Variações de nome para coluna de placa (case-insensitive)
PLATE_COLUMNS = {
    "placa",
    "placa_veiculo",
    "placa veículo",
    "placa veiculo",
    "placa do veiculo",
    "placa do veículo",
    "placa associado",
    "placa do associado",
    "placa beneficiario",
    "placa do beneficiario",
    "placa do carro",
    "placa automovel",
    "placa do automovel",
    "vehicle_plate",
    "plate",
    "veiculo_placa",
}

PLATE_ALIAS_COLUMNS = {
    "beneficio",
    "benefício",
    "beneficio associado",
    "benefício associado",
    "numero do beneficio",
    "n do beneficio",
    "número do benefício",
}

VEHICLE_MODEL_COLUMNS = {
    "modelo",
    "modelo veiculo",
    "modelo veículo",
    "modelo do veiculo",
    "modelo do veículo",
    "veiculo",
    "veículo",
    "marca modelo",
}

CPF_COLUMNS = {
    "cpf",
    "cpf cnpj",
    "cpf/cnpj",
    "documento",
    "documento associado",
}

NAME_COLUMNS = {
    "nome",
    "nome associado",
    "associado",
    "cliente",
    "nome cliente",
    "nome do associado",
}


def _normalize_plate(value: str | None) -> str | None:
    """Normaliza placa: remove não-alfanuméricos, maiúsculas. Retorna None se vazio."""
    if not value:
        return None
    normalized = re.sub(r"[^A-Za-z0-9]", "", value).upper()
    return normalized or None


def _normalize_key(value: str | None) -> str:
    """Normaliza chave de header: sem acentos, sem pontuação e com espaço único."""
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", str(value).strip().lower())
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _extract_header_candidates(row: tuple[Any, ...] | None) -> list[str]:
    if not row:
        return []
    return [str(cell).strip() for cell in row if cell not in (None, "") and str(cell).strip()]


def _detect_header_row(raw_rows: list[tuple[Any, ...]]) -> int | None:
    """Detecta a linha de cabeçalho real nas primeiras linhas da planilha."""
    best_index: int | None = None
    best_score = -1
    header_keywords = {
        "nome",
        "cpf cnpj",
        "beneficio",
        "modelo",
        "veiculo",
        "chassi",
        "situacao origem",
        "plano origem",
        "situacao destino",
        "produto destino",
    }

    for index, row in enumerate(raw_rows[:15]):
        candidates = _extract_header_candidates(row)
        if not candidates:
            continue

        normalized = {_normalize_key(cell) for cell in candidates}
        score = len(candidates)

        if _find_plate_column(normalized):
            score += 100

        score += sum(1 for item in normalized if item in header_keywords)

        if score > best_score:
            best_score = score
            best_index = index

    return best_index


def _extract_rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Extrai linhas de arquivo XLSX usando openpyxl. Pula línhas vazias."""
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        raise ValueError("Arquivo XLSX está vazio ou sem headers")

    header_row_index = _detect_header_row(raw_rows)
    if header_row_index is None:
        raise ValueError("Nenhum cabeçalho detectado no XLSX")

    headers = raw_rows[header_row_index]
    normalized_headers = [str(h).strip() if h is not None else "" for h in headers]
    if not any(normalized_headers):
        raise ValueError("Nenhum cabeçalho detectado no XLSX")

    rows: list[dict[str, Any]] = []
    for sheet_row_index, row in enumerate(raw_rows[header_row_index + 1 :], start=header_row_index + 2):
        if row is None:
            continue
        payload = {
            normalized_headers[i]: row[i] if i < len(row) else None
            for i in range(len(normalized_headers))
            if normalized_headers[i]
        }
        # Ignora linha completamente vazia
        if any(v not in (None, "") for v in payload.values()):
            payload[SOURCE_ROW_KEY] = sheet_row_index
            rows.append(payload)
    
    return rows


def _extract_rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    """Extrai linhas de arquivo CSV usando csv.DictReader."""
    buffer = StringIO(content.decode("utf-8", errors="ignore"))
    reader = csv.DictReader(buffer)
    if reader.fieldnames is None or not reader.fieldnames:
        raise ValueError("Arquivo CSV está vazio ou sem headers")
    
    rows = [dict(r) for r in reader if any(v for v in r.values() if v)]  # Ignora linhas vazias
    return rows


def _extract_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    """Dispatcher: detecta formato e extrai linhas."""
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _extract_rows_from_xlsx(content)
    if lower.endswith(".csv"):
        return _extract_rows_from_csv(content)
    raise ValueError("Formato inválido. Envie .xlsx ou .csv")


def _find_matching_column(headers: set[str], candidates: set[str], contains_terms: tuple[str, ...] = ()) -> str | None:
    normalized_candidates = {_normalize_key(candidate) for candidate in candidates}

    for key in headers:
        if key in normalized_candidates:
            return key

    if contains_terms:
        for key in headers:
            if all(term in key for term in contains_terms):
                return key

    return None


def _find_plate_column(headers: set[str]) -> str | None:
    """Encontra nome da coluna de placa nos headers normalizados."""
    explicit_plate_column = _find_matching_column(headers, PLATE_COLUMNS, ("placa",))
    if explicit_plate_column:
        return explicit_plate_column

    return _find_matching_column(headers, PLATE_ALIAS_COLUMNS, ("beneficio",))


def _find_vehicle_model_column(headers: set[str]) -> str | None:
    return _find_matching_column(headers, VEHICLE_MODEL_COLUMNS, ("modelo",))


def _find_cpf_column(headers: set[str]) -> str | None:
    cpf_column = _find_matching_column(headers, CPF_COLUMNS, ("cpf",))
    if cpf_column:
        return cpf_column

    return _find_matching_column(headers, CPF_COLUMNS, ("documento",))


def _find_name_column(headers: set[str]) -> str | None:
    name_column = _find_matching_column(headers, NAME_COLUMNS, ("nome",))
    if name_column:
        return name_column

    return _find_matching_column(headers, NAME_COLUMNS, ("associado",))


def _find_identifier_columns(headers: set[str]) -> dict[str, str | None]:
    return {
        "placa": _find_plate_column(headers),
        "modelo_veiculo": _find_vehicle_model_column(headers),
        "cpf": _find_cpf_column(headers),
        "nome": _find_name_column(headers),
    }


def _get_row_value_by_normalized_column(row: dict[str, Any], normalized_column: str | None) -> Any:
    if not normalized_column:
        return None

    for key, value in row.items():
        if key != SOURCE_ROW_KEY and _normalize_key(key) == normalized_column:
            return value

    return None


def _normalize_document(value: str | None) -> str | None:
    if not value:
        return None
    normalized = re.sub(r"\D", "", value)
    return normalized or None


def _normalize_text_identifier(value: str | None) -> str | None:
    normalized = _normalize_key(value)
    return normalized or None


def _extract_identity_fields(row: dict[str, Any], identifier_columns: dict[str, str | None]) -> dict[str, str | None]:
    plate_raw = _get_row_value_by_normalized_column(row, identifier_columns["placa"])
    model_raw = _get_row_value_by_normalized_column(row, identifier_columns["modelo_veiculo"])
    cpf_raw = _get_row_value_by_normalized_column(row, identifier_columns["cpf"])
    name_raw = _get_row_value_by_normalized_column(row, identifier_columns["nome"])

    plate = _normalize_plate(str(plate_raw) if plate_raw is not None else None)
    vehicle_model = _normalize_text_identifier(str(model_raw) if model_raw is not None else None)
    cpf = _normalize_document(str(cpf_raw) if cpf_raw is not None else None)
    name = _normalize_text_identifier(str(name_raw) if name_raw is not None else None)

    identifier_type: str | None = None
    identifier_value: str | None = None

    if plate:
        identifier_type = "placa"
        identifier_value = plate
    elif vehicle_model:
        identifier_type = "modelo_veiculo"
        identifier_value = vehicle_model
    elif cpf:
        identifier_type = "cpf"
        identifier_value = cpf
    elif name:
        identifier_type = "nome"
        identifier_value = name

    return {
        "placa_normalizada": plate,
        "modelo_veiculo": vehicle_model,
        "cpf": cpf,
        "nome": name,
        "identificador_tipo": identifier_type,
        "identificador_valor": identifier_value,
    }


def _validate_file_structure(rows: list[dict[str, Any]]) -> None:
    """Valida que arquivo tem ao menos um identificador utilizável."""
    if not rows:
        raise ValueError("Arquivo sem dados (ou apenas headers)")
    
    first_row_keys = {_normalize_key(k) for k in rows[0].keys() if k != SOURCE_ROW_KEY}
    identifier_columns = _find_identifier_columns(first_row_keys)
    
    if not any(identifier_columns.values()):
        found_headers = ", ".join(str(key) for key in rows[0].keys() if key != SOURCE_ROW_KEY)
        raise ValueError(
            "Nenhuma coluna de identificacao encontrada. "
            "A planilha precisa ter ao menos um destes campos, nesta ordem de prioridade: placa, modelo do veiculo, cpf ou nome. "
            f"Cabecalhos encontrados: {found_headers}"
        )


def _build_duplicate_reason(identifier_type: str | None, source: str) -> str:
    if identifier_type == "placa":
        return f"placa_duplicada_{source}"
    if identifier_type:
        return f"{identifier_type}_duplicado_{source}"
    return f"identificador_duplicado_{source}"


def _snapshot_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if key != SOURCE_ROW_KEY}


def _build_reference_indexes(rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, list[dict[str, Any]]]], int]:
    if not rows:
        return {"placa": {}, "modelo_veiculo": {}, "cpf": {}, "nome": {}}, 0

    first_row_normalized = {_normalize_key(k): k for k in rows[0].keys() if k != SOURCE_ROW_KEY}
    identifier_columns = _find_identifier_columns(set(first_row_normalized.keys()))
    indexes: dict[str, dict[str, list[dict[str, Any]]]] = {
        "placa": {},
        "modelo_veiculo": {},
        "cpf": {},
        "nome": {},
    }

    for row in rows:
        identity_fields = _extract_identity_fields(row, identifier_columns)
        snapshot = {
            "linha": row.get(SOURCE_ROW_KEY),
            "dados_base": _snapshot_row(row),
        }
        reference_values = {
            "placa": identity_fields.get("placa_normalizada"),
            "modelo_veiculo": identity_fields.get("modelo_veiculo"),
            "cpf": identity_fields.get("cpf"),
            "nome": identity_fields.get("nome"),
        }

        for field_name, value in reference_values.items():
            if not value:
                continue
            indexes[field_name].setdefault(value, []).append(snapshot)

    return indexes, len(rows)


def _resolve_match_against_reference(
    match: dict[str, Any],
    indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> tuple[str, str | None, dict[str, Any] | None]:
    identifiers = (
        ("placa", match.get("placa_normalizada")),
        ("modelo_veiculo", match.get("dados_interno", {}).get("modelo_veiculo")),
        ("cpf", match.get("dados_interno", {}).get("cpf")),
        ("nome", match.get("dados_interno", {}).get("nome")),
    )

    for identifier_type, identifier_value in identifiers:
        if not identifier_value:
            continue

        candidates = indexes.get(identifier_type, {}).get(identifier_value, [])
        if len(candidates) == 1:
            return "encontrado", identifier_type, candidates[0]
        if len(candidates) > 1:
            return "ambiguo", identifier_type, {"total": len(candidates)}

    return "nao_encontrado", None, None



def process_reconciliation_file(filename: str, content: bytes) -> ProcessResult:
    """
    Processa arquivo de conciliação:
    1. Extrai e valida estrutura
    2. Normaliza placas
    3. Detecta duplicatas E placas inválidas
    4. Retorna resultado para posterior cruzamento com base interna
    
    IMPORTANTE: Não faz cruzamento neste ponto. Apenas preprocessing.
    Cruzamento é feito async em conciliacoes.py com função crosscheck.
    """
    rows = _extract_rows(filename, content)
    _validate_file_structure(rows)
    
    if not rows:
        return ProcessResult(0, 0, 0, [])

    first_row_normalized = {_normalize_key(k): k for k in rows[0].keys() if k != SOURCE_ROW_KEY}
    identifier_columns = _find_identifier_columns(set(first_row_normalized.keys()))

    seen_identifiers: dict[tuple[str, str] | None, int] = {}
    matches: list[dict[str, Any]] = []
    
    for idx, row in enumerate(rows):
        identity_fields = _extract_identity_fields(row, identifier_columns)
        plate = identity_fields["placa_normalizada"]
        identifier_key = None
        if identity_fields["identificador_tipo"] and identity_fields["identificador_valor"]:
            identifier_key = (identity_fields["identificador_tipo"], identity_fields["identificador_valor"])
        
        # Marca duplicatas com fallback de identificador: placa -> modelo -> cpf -> nome.
        if identifier_key:
            seen_identifiers[identifier_key] = seen_identifiers.get(identifier_key, 0) + 1
        else:
            seen_identifiers[None] = seen_identifiers.get(None, 0) + 1
        
        row_data = _snapshot_row(row)

        # Status será determinado no cruzamento, inicialmente assume-se nao_encontrado
        status = IntegrationMatchStatus.nao_encontrado
        internal_data = {
            "linha": row.get(SOURCE_ROW_KEY, idx + 2),
            "motivo_preprocessamento": None,
            "identificador_tipo": identity_fields["identificador_tipo"],
            "identificador_valor": identity_fields["identificador_valor"],
            "modelo_veiculo": identity_fields["modelo_veiculo"],
            "cpf": identity_fields["cpf"],
            "nome": identity_fields["nome"],
        }
        
        # Validações de preprocessamento
        if not identifier_key:
            status = IntegrationMatchStatus.divergente
            internal_data["motivo_preprocessamento"] = "nenhum_identificador_localizado"
        
        matches.append({
            "placa_normalizada": plate,
            "case_id": None,
            "status_match": status,  # Preliminar; será revisado no crosscheck
            "dados_fornecedor": row_data,
            "dados_interno": internal_data,
        })

    duplicate_identifiers = {key for key, count in seen_identifiers.items() if key is not None and count > 1}
    total_divergencias = 0
    
    for match in matches:
        identifier_type = match["dados_interno"].get("identificador_tipo")
        identifier_value = match["dados_interno"].get("identificador_valor")
        identifier_key = None
        if identifier_type and identifier_value:
            identifier_key = (identifier_type, identifier_value)

        if identifier_key in duplicate_identifiers:
            match["status_match"] = IntegrationMatchStatus.divergente
            if not match["dados_interno"].get("motivo_preprocessamento"):
                match["dados_interno"]["motivo_preprocessamento"] = _build_duplicate_reason(identifier_type, "no_arquivo")
            total_divergencias += 1
        elif match["dados_interno"].get("motivo_preprocessamento"):
            total_divergencias += 1
            match["status_match"] = IntegrationMatchStatus.divergente
    
    total = len(matches)
    total_matches = total - total_divergencias
    
    return ProcessResult(
        total_registros=total,
        total_matches=total_matches,
        total_divergencias=total_divergencias,
        matches=matches,
    )


def compare_matches_with_active_report(
    matches: list[dict[str, Any]],
    filename: str,
    content: bytes,
) -> tuple[int, int, int, int]:
    rows = _extract_rows(filename, content)
    _validate_file_structure(rows)
    indexes, total_base_registros = _build_reference_indexes(rows)

    total_matches_corrected = 0
    total_divergencias_corrected = 0

    for match in matches:
        if match["dados_interno"].get("motivo_preprocessamento"):
            match["status_match"] = IntegrationMatchStatus.divergente
            total_divergencias_corrected += 1
            continue

        resolution, identifier_type, payload = _resolve_match_against_reference(match, indexes)

        if resolution == "encontrado" and payload is not None:
            match["status_match"] = IntegrationMatchStatus.encontrado
            match["dados_interno"]["motivo_preprocessamento"] = None
            match["dados_interno"]["origem_comparacao"] = "relatorio_veiculos_ativos"
            match["dados_interno"]["identificador_localizado"] = identifier_type
            match["dados_interno"]["linha_base"] = payload.get("linha")
            match["dados_interno"]["dados_base"] = payload.get("dados_base")
            total_matches_corrected += 1
            continue

        if resolution == "ambiguo":
            match["status_match"] = IntegrationMatchStatus.divergente
            match["dados_interno"]["motivo_preprocessamento"] = _build_duplicate_reason(identifier_type, "no_relatorio_base")
            match["dados_interno"]["origem_comparacao"] = "relatorio_veiculos_ativos"
            match["dados_interno"]["quantidade_correspondencias_base"] = payload.get("total") if payload else None
            total_divergencias_corrected += 1
            continue

        match["status_match"] = IntegrationMatchStatus.nao_encontrado
        match["dados_interno"]["motivo_preprocessamento"] = "nao_localizado_no_relatorio_base"
        match["dados_interno"]["origem_comparacao"] = "relatorio_veiculos_ativos"
        total_divergencias_corrected += 1

    return total_matches_corrected, total_divergencias_corrected, total_matches_corrected, total_base_registros


async def crosscheck_matches_with_database(
    matches: list[dict[str, Any]], 
    session: Any  # AsyncSession, mas evita import circular
) -> tuple[int, int, int]:
    """
    Cruza matches com a base interna (`cases` table).
    Atualiza status_match e dados_interno baseado no cruzamento.
    
    Retorna: (total_matches_corrigido, total_divergencias_corrigido, total_matches_encontrados)
    """
    total_matches_corrected = 0
    total_divergencias_corrected = 0

    for match in matches:
        if match["dados_interno"].get("motivo_preprocessamento"):
            match["status_match"] = IntegrationMatchStatus.divergente
            total_divergencias_corrected += 1
            continue

        match["status_match"] = IntegrationMatchStatus.nao_encontrado
        match["dados_interno"]["motivo_preprocessamento"] = "base_interna_nao_configurada_para_cruzamento"
        total_divergencias_corrected += 1
    
    total_encontrados = sum(1 for m in matches if m["status_match"] == IntegrationMatchStatus.encontrado)
    
    return total_matches_corrected, total_divergencias_corrected, total_encontrados